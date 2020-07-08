import openpyxl as xl
#加载excel表格
wb=xl.load_workbook('transactions.xlsx')
#得到第一张引用表
sheet = wb['Sheet1']
#访问表格
# cell = sheet['a1']
# cell = sheet.cell(1,1)
for row in range(2,sheet.max_row+1):
    cell = sheet.cell(row,3)
    correct_price= cell.value*0.9
    correct_price_cell = sheet.cell(row,4)
    correct_price_cell.value=correct_price
wb.save('transactions2.xlsx')
